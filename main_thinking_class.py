import openpyxl
import os
from datetime import datetime

class attend_thinking:
    def __init__(self):
        self.path_date = os.path.join(os.getcwd(), 'date.xlsx')
        self.path_data = os.path.join(os.getcwd(), 'data.xlsx')
        self.start_row = 1
        self.xlsx_path = ''
        self.img_path = ''
        self.name = ''

        try:
            self.date_is()
            self.row_check()
            
        except FileNotFoundError:
            self.date_init()
            self.row_check()

    def date_is(self):
        files = openpyxl.load_workbook(self.path_date)
    
    def row_check(self):
        files = openpyxl.load_workbook(self.path_date)
        dt = files.active
        while True:
            if dt.cell(row= self.start_row, column=1).value != None:
                self.start_row = self.start_row + 1
            else:
                break

    def date_init(self):
        wb = openpyxl.Workbook()
        dt = wb.active

        dt.cell(row=1, column=1).value = '이름'
        dt.cell(row=1, column=2).value = '학번'
        dt.cell(row=1, column=3).value = '시간'

        wb.save(self.path_date)

    def find_img_name(self, user):
        path = os.getcwd() + f'\\img\\20{user[0:2]}\\1-{int(user[4:6])}'
        file_list = os.listdir(path)
        for fi in file_list:
            if "학생증 명렬표" in fi and ".xlsx" in fi:
                file_name = fi
                break
        print(file_name)

        self.xlsx_path = path + "\\" + file_name
        
        files = openpyxl.load_workbook(self.xlsx_path)
        dt = files.active
        row = 2
        while True:
            if str(dt.cell(row=row,column=4).value) == user:
                self.name = dt.cell(row=row,column=3).value
                img = dt.cell(row=row,column=1).value
                break
            row = row + 1
        
        self.img_path = os.getcwd() + f'\\img\\20{user[0:2]}\\1-{int(user[4:6])}\\{img}.jpg'


    def date_add(self, user):
        files = openpyxl.load_workbook(self.path_date)
        dt = files.active

        dt.cell(row=self.start_row, column=1).value = self.name
        dt.cell(row=self.start_row, column=2).value = user
        dt.cell(row=self.start_row, column=3).value = str(datetime.now())
        self.start_row = self.start_row + 1

        files.save(self.path_date)

# att = attend_thinking()

# att.find_img_name("19010402")
# print(att.img_path, att.name)
# user = "19010402"
# print(f"{int(user[0:2])}\\{int(user[4:6])}\\{int(user[6:8])}")